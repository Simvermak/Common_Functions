import json
import pandas as pd
from datetime import datetime, timedelta
from sqlalchemy import text
import openpyxl

import sys

sys.path.append("..")
import common.database as database
import ExcelTemplate as xlst


def main():
    '''
    说明：客户状况汇总（月）
    状态：
    '''
    print('客户状况汇总（月）')
    today = datetime.today()
    yesterday = today - timedelta(days=1)
    yesterday_str = yesterday.strftime(r'%Y-%m-%d')

    engine = database.engine('ICCPP_DW_DWS')

    # 客户信息
    sql = f'''
    SELECT
      a.customerName,
      a.ownerUserName,
      a.region3,
      a.region4,
      a.crm_code,
      a.bos_code,
      b.last_month_order_amount,
      b.month_order_amount,
      b.last_month_shipping_amount,
      b.month_shipping_amount
    FROM
      ICCPP_DW_DWS.crm_customer a
    left join 
          bos_customer_state b
          on
      a.bos_code = b.bos_code
    WHERE
      a.region2 = '海外营销中心'
      and a.region3 is not null
      and a.region4 is not null
    '''
    df_custom = pd.read_sql(text(sql), engine.connect())

    # 拜访信息
    sql = f'''
    SELECT
      a.targetIntro,
      a.crm_code,
      b.value as visit_date
    FROM
      ICCPP_DW_DWS.crm_form a
    left join ICCPP_DW_DWS.crm_form_data b on
      a.form_id = b.form_id
      and name = 'Date of Visit'
    WHERE
      a.form_key IN ('7ErjR9Pu', 'epo32S8g')
      and a.createTime<DATE_FORMAT(NOW(), '%Y-%m-%d')
    '''
    df_visit = pd.read_sql(text(sql), engine.connect())

    # 获取昨天所在月份
    current_month = yesterday.strftime("%Y-%m")
    # 获取昨天上个月份
    prev_month = (yesterday.replace(day=1) - timedelta(days=1)).strftime('%Y-%m')

    # 当月拜访
    df_visit_this_month = df_visit[df_visit['visit_date'].str.contains(current_month, na=False)]
    # 上月拜访
    df_visit_last_month = df_visit[df_visit['visit_date'].str.contains(prev_month, na=False)]
    # 当月拜访次数
    df_visit_this_month_num = df_visit_this_month['crm_code'].value_counts().reset_index()
    df_visit_this_month_num.columns = ['crm_code', 'this_month_visit_nums']
    # 上月拜访次数
    df_visit_last_month_num = df_visit_last_month['crm_code'].value_counts().reset_index()
    df_visit_last_month_num.columns = ['crm_code', 'last_month_visit_nums']

    # 合并拜访信息
    df = pd.merge(df_custom, df_visit_this_month_num, how='left', on='crm_code')
    df = pd.merge(df, df_visit_last_month_num, how='left', on='crm_code')

    # 处理缺省值
    df['last_month_order_amount'].fillna(0, inplace=True)
    df['month_order_amount'].fillna(0, inplace=True)
    df['last_month_shipping_amount'].fillna(0, inplace=True)
    df['month_shipping_amount'].fillna(0, inplace=True)

    df['last_month_visit_nums'].fillna(0, inplace=True)
    df['this_month_visit_nums'].fillna(0, inplace=True)

    # 指定排序顺序
    custom_order = ['北美大区', '欧洲大区']
    # 按指定顺序排序
    df['order'] = df['region3'].map({v: i for i, v in enumerate(custom_order)})
    df = df.sort_values(by=['order', 'region4', 'ownerUserName', 'month_order_amount'],
                        ascending=[True, True, True, False]).drop('order', axis=1)

    file_name = '客户状况汇总（月）'
    file_path = f'report/crm/{file_name}{yesterday_str}.xlsx'

    # 传递给excel
    lo_infos = {}

    # 工作簿名字
    lo_infos['sheet_name'] = file_name

    # Rows
    lo_infos['voopooRows'] = json.loads(df.to_json(orient='records'))

    # 数据时间
    lo_infos['data_time'] = today.strftime(r'%Y-%m-%d %H:%M:%S')
    # 报告日期
    lo_infos['report_date'] = yesterday_str
    lo_infos['report_year'] = yesterday.year
    lo_infos['report_month'] = yesterday.month

    xlst.write2(lo_infos, file_path, 'CRM-客户状况汇总（月）', False)
    print('生成', file_path)

    wb = openpyxl.load_workbook(file_path)
    sheet = wb[file_name]

    # 先把所有内容放入一个List中
    sumList = []
    # 第几行开始
    start_row = 3
    end_row = start_row + sheet.max_row
    column = 4

    for i in range(start_row, end_row):
        value = sheet.cell(row=i, column=column).value
        if value:
            sumList.append(value)
        else:
            break

    # # 开始合并单元格
    # preRow = 0
    # finRow = 0
    # flag = sumList[0]
    # for i in range(len(sumList)):
    #     if sumList[i] != flag:
    #         flag = sumList[i]
    #         finRow = i - 1
    #         if finRow >= preRow:
    #             sheet.merge_cells("D{}:D{}".format(preRow+start_row, finRow+start_row))
    #             sheet.merge_cells("C{}:C{}".format(preRow+start_row, finRow+start_row))
    #             sheet.merge_cells("B{}:B{}".format(preRow+start_row, finRow+start_row))
    #             preRow = finRow + 1
    #     if i == len(sumList) - 1:
    #         finRow = i
    #         sheet.merge_cells("D{}:D{}".format(preRow+start_row, finRow+start_row))
    #         sheet.merge_cells("C{}:C{}".format(preRow+start_row, finRow+start_row))
    #         sheet.merge_cells("B{}:B{}".format(preRow+start_row, finRow+start_row))
    # # 保存表
    # wb.save(file_path)

    # wb = openpyxl.load_workbook(file_path)
    # sheet = wb["sheet"]

    # # 先把所有内容放入一个List中
    # sumList = []
    # # 第几行开始
    # start_row=3
    # end_row=start_row+sheet.max_row
    # column=3

    # for i in range(start_row, end_row):
    #     value = sheet.cell(row=i, column=column).value
    #     if value:
    #         sumList.append(value)
    #     else:
    #         break

    # # 开始合并单元格
    # preRow = 0
    # finRow = 0
    # flag = sumList[0]
    # for i in range(len(sumList)):
    #     if sumList[i] != flag:
    #         flag = sumList[i]
    #         finRow = i - 1
    #         if finRow >= preRow:
    #             sheet.merge_cells("C{}:C{}".format(preRow+start_row, finRow+start_row))
    #             sheet.merge_cells("B{}:B{}".format(preRow+start_row, finRow+start_row))
    #             preRow = finRow + 1
    #     if i == len(sumList) - 1:
    #         finRow = i
    #         sheet.merge_cells("C{}:C{}".format(preRow+start_row, finRow+start_row))
    #         sheet.merge_cells("B{}:B{}".format(preRow+start_row, finRow+start_row))
    # # 保存表
    # wb.save(file_path)

    # # 转换成pdf
    # new_path = f'report/quality/产品品质问题数据报告{yesterday_str}.pdf'
    # password=''
    # excel=convert.excel(file_path,new_path,password)
    # excel.to_pdf()  
    # print()

    print()
    return file_path


if __name__ == '__main__':
    main()
