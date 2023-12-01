import json
import pandas as pd
from datetime import datetime, timedelta
from sqlalchemy import text
import openpyxl
import calendar

import sys

sys.path.append("..")
import common.database as database
import ExcelTemplate as xlst


def main():
    '''
    说明：经营运营专项数据指标（日报）
    状态：
    '''

    today = datetime.today()
    yesterday = today - timedelta(days=1)
    yesterday_str = yesterday.strftime(r'%Y-%m-%d')

    date_range = [(today - timedelta(days=14)) + timedelta(days=i) for i in range(14)]
    date_range_int = [int(date.strftime('%Y%m%d')) for date in date_range]

    engine = database.engine('ICCPP_DW_DWS')

    # 接单数据
    sql = '''
        SELECT
            bs_so.Order_Date AS Date,
            bs_so. Business_Group_Name AS Business_Group,
            SUM(bs_so.Confimred_Order_Quantity) AS Total_Quantity,
            SUM(bs_so.Final_Order_Rebate_Amount_Cny) AS Total_Amount,
            pl.product_categories
        FROM
            ICCPP_DW_DWS.bos_so_correct bs_so
        JOIN
            ICCPP_DW_DIM.product_last pl
        ON
            bs_so.Product_Code = pl.product_code
        WHERE
            bs_so.Order_Date >= DATE_SUB(CURDATE(), INTERVAL 14 DAY)
            AND bs_so.Order_Date < CURDATE()
            AND bs_so.Business_Group_Name IN ('VOOPOO海外', 'ODM事业部', 'ZOVOO海外')
        GROUP BY
            bs_so.Order_Date,
            bs_so. Business_Group_Name,
            pl.product_categories
        ORDER BY
            bs_so.Order_Date DESC,
            bs_so.Business_Group_Name DESC,
            pl.product_categories DESC
        '''

    so_df = pd.read_sql(text(sql), engine.connect())

    # 交付数据
    sql = '''
    SELECT
        bs_sa.Stock_Out_Date AS Date,
        bs_sa. Business_Group,
        SUM(bs_sa.Stock_Out_Quantity) AS Total_Quantity,
        SUM(bs_sa.Rebate_Sales_Amount_Cny) AS Total_Amount,
        pl.product_categories
    FROM
        ICCPP_DW_DWS.bos_sa_correct bs_sa
    JOIN
        ICCPP_DW_DIM.product_last pl
    ON
        bs_sa.Product_Code = pl.product_code
    WHERE
        bs_sa.Stock_Out_Date >= DATE_SUB(CURDATE(), INTERVAL 14 DAY)
        AND bs_sa.Stock_Out_Date < CURDATE()
        AND bs_sa.Business_Group IN ('VOOPOO海外', 'ODM事业部', 'ZOVOO海外')
    GROUP BY
        bs_sa.Stock_Out_Date,
        bs_sa. Business_Group,
        pl.product_categories
    ORDER BY
        bs_sa.Stock_Out_Date DESC,
        bs_sa.Business_Group DESC,
        pl.product_categories DESC
    '''

    sa_df = pd.read_sql(text(sql), engine.connect())

    # 未交付数据
    sql = '''
    SELECT
        bs_un.undelivered_date AS Date,
        bs_un. Business_Group_Name AS Business_Group,
        SUM(bs_un.Undelivered_Quantity) AS Total_Quantity,
        SUM(bs_un.Undelivered_Order_Rebate_Amount_Cny) AS Total_Amount,
        pl.product_categories
    FROM
        ICCPP_DW_DWD.bos_undelivered bs_un
    JOIN
        ICCPP_DW_DIM.product_last pl
    ON
        bs_un.Product_Code = pl.product_code
    WHERE
        bs_un.undelivered_date >= DATE_SUB(CURDATE(), INTERVAL 14 DAY)
        AND bs_un.undelivered_date < CURDATE()
        AND bs_un.Business_Group_Name IN ('VOOPOO海外', 'ODM事业部', 'ZOVOO海外')
    GROUP BY
        bs_un.undelivered_date,
        bs_un. Business_Group_Name,
        pl.product_categories
    ORDER BY
        bs_un.undelivered_date DESC,
        bs_un.Business_Group_Name DESC,
        pl.product_categories DESC
    '''

    un_df = pd.read_sql(text(sql), engine.connect())

    # 忽略CBD的数据
    un_df = un_df[un_df['product_categories'] != 'CBD']

    sql = '''
    SELECT
        ReportDate AS Date,
        MA0,
        MA1,
        L1Subject1,
        L2Subject2,
        L3Subject3,
        D1
    FROM
        ICCPP_DW_DWS.oa_form
    WHERE
        form_type = '日报'
        AND ReportDate >= DATE_SUB(CURDATE(), INTERVAL 14 DAY)
    	AND ReportDate < CURDATE()
    '''

    df = pd.read_sql(text(sql), engine.connect())

    df['Date'] = df['Date'].str.replace('-', '').astype(int)
    df['MA0'] = df['MA0'].astype(float)
    df['MA1'] = df['MA1'].astype(float)

    # 传递给excel
    lo_infos = {}

    date_list = []
    for date in date_range:
        day_of_week = calendar.weekday(date.year, date.month, date.day)
        date_str = f'{date.month}月{date.day}号'
        if day_of_week == 5:
            date_str = '[[BG:G]]' + date_str
        elif day_of_week == 6:
            date_str = '[[BG:E]]' + date_str
        date_list.append(date_str)

    # 标题系统日期
    date_df = pd.DataFrame({'Date': date_list})
    lo_infos['ReportDate'] = date_df.to_dict(orient='records')

    def filter_data(target_df, condition=None):  # 需要将df中的判断日期字段列名改为 Date ,格式为int
        if condition is None:
            filtered_data = target_df.copy()
        else:
            filtered_data = target_df[condition].copy()

        if filtered_data.empty:
            filtered_data['Date'] = []  # 添加空的'Date'列

        # 获取在 date_range_int 中存在但在 Date 列中不存在的日期
        missing_dates = set(date_range_int) - set(filtered_data['Date'])

        # 生成缺失日期的新行
        missing_rows = pd.DataFrame(
            {'Date': list(missing_dates), **{key: 0 for key in target_df.columns if key != 'Date'}})
        new_df = pd.concat([filtered_data, missing_rows]).sort_values('Date').reset_index(drop=True)
        new_df_json = new_df.to_json(orient='records')
        filtered_data_list = json.loads(new_df_json)

        return filtered_data_list

    # 接单金额
    so_oa_df = so_df.groupby(["Date", "Business_Group"]).agg({"Total_Amount": "sum"}).reset_index()
    lo_infos['oa_voopoo'] = filter_data(so_oa_df, so_oa_df['Business_Group'] == 'VOOPOO海外')
    lo_infos['oa_zovoo'] = filter_data(so_oa_df, so_oa_df['Business_Group'] == 'ZOVOO海外')
    lo_infos['oa_odm'] = filter_data(so_oa_df, so_oa_df['Business_Group'] == 'ODM事业部')
    oa_sum = so_df.groupby('Date')['Total_Amount'].sum().reset_index()
    lo_infos['oa_sum'] = filter_data(oa_sum)

    # 接单数量
    lo_infos['oq_host_voopoo'] = filter_data(so_df, (so_df['Business_Group'] == 'VOOPOO海外') & (
            so_df['product_categories'] == '主机'))
    lo_infos['oq_part_voopoo'] = filter_data(so_df, (so_df['Business_Group'] == 'VOOPOO海外') & (
            so_df['product_categories'] == '耗材'))
    lo_infos['oq_host_zovoo'] = filter_data(so_df, (so_df['Business_Group'] == 'ZOVOO海外') & (
            so_df['product_categories'] == '主机'))
    lo_infos['oq_part_zovoo'] = filter_data(so_df, (so_df['Business_Group'] == 'ZOVOO海外') & (
            so_df['product_categories'] == '耗材'))
    lo_infos['oq_host_odm'] = filter_data(so_df, (so_df['Business_Group'] == 'ODM事业部') & (
            so_df['product_categories'] == '主机'))
    lo_infos['oq_part_odm'] = filter_data(so_df, (so_df['Business_Group'] == 'ODM事业部') & (
            so_df['product_categories'] == '耗材'))
    oq_sum = so_df.groupby('Date')['Total_Quantity'].sum().reset_index()
    lo_infos['oq_sum'] = filter_data(oq_sum)

    # 交付金额
    sa_sa_df = sa_df.groupby(["Date", "Business_Group"]).agg({"Total_Amount": "sum"}).reset_index()

    lo_infos['sa_voopoo'] = filter_data(sa_sa_df, sa_sa_df['Business_Group'] == 'VOOPOO海外')
    lo_infos['sa_zovoo'] = filter_data(sa_sa_df, sa_sa_df['Business_Group'] == 'ZOVOO海外')
    lo_infos['sa_odm'] = filter_data(sa_sa_df, sa_sa_df['Business_Group'] == 'ODM事业部')
    sa_sum = sa_df.groupby('Date')['Total_Amount'].sum().reset_index()
    lo_infos['sa_sum'] = filter_data(sa_sum)

    # 交付数量
    lo_infos['sq_host_voopoo'] = filter_data(sa_df, (sa_df['Business_Group'] == 'VOOPOO海外') & (
            sa_df['product_categories'] == '主机'))
    lo_infos['sq_part_voopoo'] = filter_data(sa_df, (sa_df['Business_Group'] == 'VOOPOO海外') & (
            sa_df['product_categories'] == '耗材'))
    lo_infos['sq_host_zovoo'] = filter_data(sa_df, (sa_df['Business_Group'] == 'ZOVOO海外') & (
            sa_df['product_categories'] == '主机'))
    lo_infos['sq_part_zovoo'] = filter_data(sa_df, (sa_df['Business_Group'] == 'ZOVOO海外') & (
            sa_df['product_categories'] == '耗材'))
    lo_infos['sq_host_odm'] = filter_data(sa_df, (sa_df['Business_Group'] == 'ODM事业部') & (
            sa_df['product_categories'] == '主机'))
    lo_infos['sq_part_odm'] = filter_data(sa_df, (sa_df['Business_Group'] == 'ODM事业部') & (
            sa_df['product_categories'] == '耗材'))

    sq_sum = sa_df.groupby('Date')['Total_Quantity'].sum().reset_index()
    lo_infos['sq_sum'] = filter_data(sq_sum)

    # 未交金额
    un_oa_df = un_df.groupby(["Date", "Business_Group"]).agg({"Total_Amount": "sum"}).reset_index()
    lo_infos['un_voopoo'] = filter_data(un_oa_df, un_oa_df['Business_Group'] == 'VOOPOO海外')
    lo_infos['un_zovoo'] = filter_data(un_oa_df, un_oa_df['Business_Group'] == 'ZOVOO海外')
    lo_infos['un_odm'] = filter_data(un_oa_df, un_oa_df['Business_Group'] == 'ODM事业部')
    ua_sum = un_oa_df.groupby('Date')['Total_Amount'].sum().reset_index()
    lo_infos['ua_sum'] = filter_data(ua_sum)

    # 未交数量
    lo_infos['un_host_voopoo'] = filter_data(un_df, (un_df['Business_Group'] == 'VOOPOO海外') & (
            un_df['product_categories'] == '主机'))
    lo_infos['un_part_voopoo'] = filter_data(un_df, (un_df['Business_Group'] == 'VOOPOO海外') & (
            un_df['product_categories'] == '耗材'))
    lo_infos['un_host_zovoo'] = filter_data(un_df, (un_df['Business_Group'] == 'ZOVOO海外') & (
            un_df['product_categories'] == '主机'))
    lo_infos['un_part_zovoo'] = filter_data(un_df, (un_df['Business_Group'] == 'ZOVOO海外') & (
            un_df['product_categories'] == '耗材'))
    lo_infos['un_host_odm'] = filter_data(un_df, (un_df['Business_Group'] == 'ODM事业部') & (
            un_df['product_categories'] == '主机'))
    lo_infos['un_part_odm'] = filter_data(un_df, (un_df['Business_Group'] == 'ODM事业部') & (
            un_df['product_categories'] == '耗材'))
    uq_sum = un_df.groupby('Date')['Total_Quantity'].sum().reset_index()
    lo_infos['uq_sum'] = filter_data(uq_sum)

    # 日产值
    df_rcz = df[df['L2Subject2'] == '日产值'].copy()

    # 产品线
    df_cpx = df_rcz[df_rcz['L3Subject3'] == 'by产品线'].copy()

    lo_infos['cpx_voopoo'] = filter_data(df_cpx, df_cpx['D1'] == 'VOOPOO')
    lo_infos['cpx_zovoo'] = filter_data(df_cpx, df_cpx['D1'] == 'ZOVOO')
    lo_infos['cpx_odm'] = filter_data(df_cpx, df_cpx['D1'] == 'ODM')

    cpx_sum = df_cpx.groupby('Date')['MA0'].sum().reset_index()
    lo_infos['cpx_sum'] = filter_data(cpx_sum)

    # 工厂
    df_gc = df[df['L3Subject3'] == 'by工厂'].copy()

    lo_infos['gc_sg'] = filter_data(df_gc, df_gc['D1'] == '松岗')
    lo_infos['gc_lb'] = filter_data(df_gc, df_gc['D1'] == '寮步')
    lo_infos['gc_dls'] = filter_data(df_gc, df_gc['D1'] == '大岭山')
    lo_infos['gc_wx'] = filter_data(df_gc, df_gc['D1'] == '外协')

    gc_sum = df_gc.groupby('Date')['MA0'].sum().reset_index()
    lo_infos['gc_sum'] = filter_data(gc_sum)

    # 人均产值
    df_rjcz = df[df['L2Subject2'] == '人均产值'].copy()

    lo_infos['rjcz_sg'] = filter_data(df_rjcz, df_rjcz['D1'] == '松岗')
    lo_infos['rjcz_lb'] = filter_data(df_rjcz, df_rjcz['D1'] == '寮步')
    lo_infos['rjcz_dls'] = filter_data(df_rjcz, df_rjcz['D1'] == '大岭山')

    rjcz_sum0 = df_rjcz.groupby('Date')['MA0'].sum().reset_index()
    rjcz_sum1 = df_rjcz.groupby('Date')['MA1'].sum().reset_index()
    lo_infos['rjcz_sum0'] = filter_data(rjcz_sum0)
    lo_infos['rjcz_sum1'] = filter_data(rjcz_sum1)

    # 工厂人数
    df_gcrs = df[df['L2Subject2'] == '工厂人数'].copy()

    lo_infos['gcrs_sg'] = filter_data(df_gcrs, df_gcrs['D1'] == '松岗')
    lo_infos['gcrs_lb'] = filter_data(df_gcrs, df_gcrs['D1'] == '寮步')
    lo_infos['gcrs_dls'] = filter_data(df_gcrs, df_gcrs['D1'] == '大岭山')
    gcrs_sum = df_gcrs.groupby('Date').agg({'MA0': 'sum', 'MA1': 'sum'}).reset_index()
    lo_infos['gcrs_sum'] = filter_data(gcrs_sum)

    # 综合生产效率
    df_scxl = df[df['L2Subject2'] == '综合生产效率'].copy()

    lo_infos['scxl_sg'] = filter_data(df_scxl, df_scxl['D1'] == '松岗')
    lo_infos['scxl_lb'] = filter_data(df_scxl, df_scxl['D1'] == '寮步')
    lo_infos['scxl_dls'] = filter_data(df_scxl, df_scxl['D1'] == '大岭山')

    # 损失工时
    df_ssgs = df[df['L2Subject2'] == '损失工时'].copy()

    lo_infos['ssgs_sg'] = filter_data(df_ssgs, df_ssgs['D1'] == '松岗')
    lo_infos['ssgs_lb'] = filter_data(df_ssgs, df_ssgs['D1'] == '寮步')
    lo_infos['ssgs_dls'] = filter_data(df_ssgs, df_ssgs['D1'] == '大岭山')

    # 工厂稼动率
    df_gcjdl = df[df['L2Subject2'] == '工厂稼动率'].copy()

    lo_infos['gcjdl_sg'] = filter_data(df_gcjdl, df_gcjdl['D1'] == '松岗')
    lo_infos['gcjdl_lb'] = filter_data(df_gcjdl, df_gcjdl['D1'] == '寮步')
    lo_infos['gcjdl_dls'] = filter_data(df_gcjdl, df_gcjdl['D1'] == '大岭山')

    # 12H物料齐套率
    df_wlqtl = df[df['L2Subject2'] == '12H物料齐套率'].copy()

    lo_infos['wlqtl_sg'] = filter_data(df_wlqtl, df_wlqtl['D1'] == '松岗')
    lo_infos['wlqtl_lb'] = filter_data(df_wlqtl, df_wlqtl['D1'] == '寮步')
    lo_infos['wlqtl_dls'] = filter_data(df_wlqtl, df_wlqtl['D1'] == '大岭山')

    # 品质异常关闭率
    df_pzycgbl = df[df['L2Subject2'] == '品质异常关闭率'].copy()

    lo_infos['pzycgbl_sg'] = filter_data(df_pzycgbl, df_pzycgbl['D1'] == '松岗')
    lo_infos['pzycgbl_lb'] = filter_data(df_pzycgbl, df_pzycgbl['D1'] == '寮步')
    lo_infos['pzycgbl_dls'] = filter_data(df_pzycgbl, df_pzycgbl['D1'] == '大岭山')

    # 72小时内计划变动率
    df_jhbdl = df[df['L2Subject2'] == '72小时内计划变动率'].copy()

    lo_infos['jhbdl_sg'] = filter_data(df_jhbdl, df_jhbdl['D1'] == '松岗')
    lo_infos['jhbdl_lb'] = filter_data(df_jhbdl, df_jhbdl['D1'] == '寮步')
    lo_infos['jhbdl_dls'] = filter_data(df_jhbdl, df_jhbdl['D1'] == '大岭山')

    # IQC来料检验合格率-量产
    df_iqc = df[df['L2Subject2'] == 'IQC来料检验合格率-量产'].copy()

    lo_infos['iqc_sg'] = filter_data(df_iqc, df_iqc['D1'] == '松岗')
    lo_infos['iqc_lb'] = filter_data(df_iqc, df_iqc['D1'] == '寮步')
    lo_infos['iqc_dls'] = filter_data(df_iqc, df_iqc['D1'] == '大岭山')

    # 盲点测试合格率
    df_mdcshgl = df[df['L2Subject2'] == '盲点测试合格率'].copy()

    lo_infos['mdcshgl_sg'] = filter_data(df_mdcshgl, df_mdcshgl['D1'] == '松岗')
    lo_infos['mdcshgl_lb'] = filter_data(df_mdcshgl, df_mdcshgl['D1'] == '寮步')
    lo_infos['mdcshgl_dls'] = filter_data(df_mdcshgl, df_mdcshgl['D1'] == '大岭山')

    # 制程合格率
    df_zchgl = df[df['L2Subject2'] == '制程合格率'].copy()

    lo_infos['zchgl_sg'] = filter_data(df_zchgl, df_zchgl['D1'] == '松岗')
    lo_infos['zchgl_lb'] = filter_data(df_zchgl, df_zchgl['D1'] == '寮步')
    lo_infos['zchgl_dls'] = filter_data(df_zchgl, df_zchgl['D1'] == '大岭山')

    # FQC批次合格率
    df_fqc = df[df['L2Subject2'] == 'FQC批次合格率'].copy()

    lo_infos['fqc_sg'] = filter_data(df_fqc, df_fqc['D1'] == '松岗')
    lo_infos['fqc_lb'] = filter_data(df_fqc, df_fqc['D1'] == '寮步')
    lo_infos['fqc_dls'] = filter_data(df_fqc, df_fqc['D1'] == '大岭山')

    # lo_infos['soRows'] = json.loads(df_project.to_json(orient='records'))

    lo_infos['sheet_name'] = 'sheet'

    # 数据时间
    lo_infos['data_time'] = today.strftime(r'%Y-%m-%d %H:%M:%S')
    # 报告日期
    lo_infos['report_date'] = yesterday_str

    file_path = f'report/operate/经营运营专项数据指标 日报告 {yesterday_str}.xlsx'

    xlst.write2(lo_infos, file_path, '经营运营专项数据指标 日报告', True)

    print('生成', file_path)

    # 加载现有工作簿
    workbook = openpyxl.load_workbook(file_path)
    # 选择工作表
    worksheet = workbook.active
    # 指定要删除的特定内容
    target_content = '((del))'
    # 遍历所有列并检查第一行是否包含指定内容
    for column in worksheet.iter_cols():
        first_cell = column[2].value

        if first_cell and target_content in first_cell:
            # 如果包含指定内容，则删除该列
            worksheet.delete_cols(column[2].column)

    workbook.save(file_path)

    return file_path  

    exit()

    # 打开Excel文件
    workbook = openpyxl.load_workbook(file_path)

    # 遍历所有工作表
    for sheet_name in workbook.sheetnames:

        worksheet = workbook[sheet_name]
        # for column in worksheet.columns:
        #   column_index = column[0].column_letter
        #   if worksheet.column_dimensions[column_index].hidden:
        #       print(column_index)
        #       worksheet.delete_cols(column_index)  # 使用delete_cols()方法删除列

        for column in worksheet.columns:
            column_index = column[0].column_letter
            if worksheet.column_dimensions[column_index].hidden:
                worksheet.delete_column(column_index)

                # 保存修改后的Excel文件
    workbook.save(file_path)

    return file_path


if __name__ == '__main__':
    main()
