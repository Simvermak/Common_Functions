# -*- coding: utf-8 -*-
import os
import re
import datetime
from time import strftime, localtime
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from xltpl.writerx import BookWriter # 需pip install xltpl
'''
https://github.com/zhangyu836/python-xlsx-template
'''

# 淡绿色
fillA = PatternFill(start_color='00C6E0B4', end_color='00C6E0B4', fill_type="solid")
# 更淡绿色
fillB = PatternFill(start_color='00DEEDD4', end_color='00DEEDD4', fill_type="solid")
# 淡黄色
fillC = PatternFill(start_color='00FFEDB3', end_color='00FFEDB3', fill_type="solid")
# 淡橙色
fillD = PatternFill(start_color='00F5B48D', end_color='00F5B48D', fill_type="solid")
# 深橙色
fillE = PatternFill(start_color='00EF894B', end_color='00EF894B', fill_type="solid")
# 蓝色
fillF = PatternFill(start_color='00C7ECFF', end_color='00C7ECFF', fill_type="solid")
# 绿色
fillG = PatternFill(start_color='0070AD47', end_color='0070AD47', fill_type="solid")
# 橙红色
fillH = PatternFill(start_color='FFFF2400', end_color='FFFF2400', fill_type="solid")
# 黄色
fillI = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type="solid")

# 绿色
lightA = Font(size=18, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='0070AD47')
# 淡绿色
lightB = Font(size=18, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='00B1D597')
# 黄色
lightC = Font(size=18, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='00FFDF29')
# 黄橙色
lightD = Font(size=18, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='00F4AB7B')
# 深橙色
lightE = Font(size=18, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='00ED7D31')

lightA2 = Font(bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='00165b13',name='Arial')
lightE2 = Font(bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='00ED7D31',name='Arial')


def write(lo_infos,file_name,tmpl_name,is_stain=True):
  pth = os.path.dirname(__file__)
  resultFname = os.path.join(pth, file_name)
  payloads=[]

  templateFname = os.path.join(pth, f'template/{tmpl_name}.xlsx')
  writer = BookWriter(templateFname)
  writer.jinja_env.globals.update(dir=dir, getattr=getattr)

  if isinstance(lo_infos, list):
      # 多工作簿模式
      payloads = lo_infos
  else:
      payloads.append(lo_infos)

  writer.render_book(payloads=payloads)
  writer.save(resultFname)
  
  if is_stain:
    # 二次读取，遍历单元格，通过单元格样式代码，对单元格进行样式标记
    wb = load_workbook(filename=resultFname)
    for ws in wb:  # 遍历sheet
        for row in ws.rows:  # 遍历行

            row_number = row[0].row

            for cell in row:  # 遍列
              if isinstance(cell.value,str):  # 列不为空时，读取例数据
                  #print(cell.value, cell.row, cell.column, cell.coordinate)
                  matchObj = re.search(r'\[\[.*]]', cell.value, re.I)  # 正则判断列中数否有样式代码
                  if matchObj:  # 如果有样式代码，则进行样式风格渲染
                      cellStyle = matchObj.group(0)  # 通过正则获得样式代码
                      cell.value = cell.value.replace(cellStyle, "")  # 抹掉字段值中的样式代码
                      match cellStyle:  # 判断样式代码的类型
                          # 字体
                          case "[[FT:A]]":
                              ws[cell.coordinate].font = lightA
                          case "[[FT:B]]":
                              ws[cell.coordinate].font = lightB
                          case "[[FT:C]]":
                              ws[cell.coordinate].font = lightC
                          case "[[FT:D]]":
                              ws[cell.coordinate].font = lightD
                          case "[[FT:E]]":
                              ws[cell.coordinate].font = lightE
                          case "[[FT:A2]]":
                              ws[cell.coordinate].font = lightA2
                          case "[[FT:E2]]":
                              ws[cell.coordinate].font = lightE2        

                          # 背景
                          case "[[BG:A]]":
                              ws[cell.coordinate].fill = fillA
                          case "[[BG:B]]":
                              ws[cell.coordinate].fill = fillB
                          case "[[BG:C]]":
                              ws[cell.coordinate].fill = fillC
                          case "[[BG:D]]":
                              ws[cell.coordinate].fill = fillD
                          case "[[BG:E]]":
                              ws[cell.coordinate].fill = fillE
                          case "[[BG:F]]":
                              ws[cell.coordinate].fill = fillF   
                          case "[[BG:G]]":
                              ws[cell.coordinate].fill = fillG
                          case "[[BG:H]]":
                              ws[cell.coordinate].fill = fillH
                          case "[[BG:I]]":
                              ws[cell.coordinate].fill = fillI

                          # 自动行高
                          case "[[HEIGHT:AUTO]]":
                              count = cell.value.count('\n')
                              row_dimensions = ws.row_dimensions[row_number]
                              row_dimensions.height = 20*(1+count)
                  
                  # 填充空值    
                  if cell.value=='None':
                      cell.value = '--'        

    wb.save(filename=resultFname)