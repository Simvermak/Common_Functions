import time
# 解决 ------------------------
#  main thread is not in main loop
import matplotlib
matplotlib.use('Agg')
# 解决 ------------------------
import matplotlib.pyplot as plt

import json
from apitable import Apitable
import pandas as pd
from datetime import datetime, timedelta
from sqlalchemy import text
import numpy as np
import os
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Alignment, Border, Side

import sys
sys.path.append("..")
import common.convert as convert
import common.database as database
import ExcelTemplate as xlst

def main():
    '''
    说明：客户报告
    状态：
    '''
        



    today = datetime.today()
    yesterday = today - timedelta(days=1)
    yesterday_str = yesterday.strftime(r'%Y-%m-%d')

    engine = database.engine('ICCPP_DW_DWS')

    # 传递给excel
    lo_infos = {}

    # 客户信息
    sql=f'''
    SELECT
      bos_code as BOS_Group_Id,
      name as Customer_Name,
      crm_code,
      sale_org_name as Business_Group_Name,
      sale_area_name as Business_Region
    FROM
      ICCPP_DW_DWS.bos_customer_state
    WHERE
      isactive = '是'
    '''
    df_custom = pd.read_sql(text(sql), engine.connect())

    # 上月接单
    sql=f'''
    SELECT
      BOS_Group_Id,
      Sales_Representative_Name,
      sum(Final_Order_Rebate_Amount_Cny) as last_month_order_amount,
      MAX(Order_Date) as Order_Date
    FROM
      ICCPP_DW_DWS.bos_so_correct
    WHERE
      DATE_FORMAT(Order_Date, '%Y-%m') = DATE_FORMAT(NOW() - INTERVAL 1 MONTH, '%Y-%m')
    group by
      BOS_Group_Id,
      Sales_Representative_Name
    order by
      BOS_Group_Id,
      Order_Date
    '''
    df_so_last_month = pd.read_sql(text(sql), engine.connect())
    df_so_last_month=df_so_last_month.groupby(['BOS_Group_Id']).agg({'last_month_order_amount':'sum','Sales_Representative_Name':'last'}).reset_index()


    # 本月接单
    sql=f'''
    SELECT
      BOS_Group_Id,
      Sales_Representative_Name,
      sum(Final_Order_Rebate_Amount_Cny) as this_month_order_amount,
      MAX(Order_Date) as Order_Date
    FROM
      ICCPP_DW_DWS.bos_so_correct
    WHERE
      DATE_FORMAT(Order_Date, '%Y-%m') = DATE_FORMAT(NOW(), '%Y-%m')
    group by
      BOS_Group_Id,
      Sales_Representative_Name
    order by
      BOS_Group_Id,
      Order_Date  
    '''
    df_so_this_month = pd.read_sql(text(sql), engine.connect())
    df_so_this_month=df_so_this_month.groupby(['BOS_Group_Id']).agg({'this_month_order_amount':'sum','Sales_Representative_Name':'last'}).reset_index()
    
    # 上月出货
    sql=f'''
    SELECT
      BOS_Group_Id,
      sum(Rebate_Sales_Amount_Cny) as last_month_sales_amount
    FROM
      ICCPP_DW_DWS.bos_sa_correct
    WHERE
      DATE_FORMAT(Stock_Out_Date, '%Y-%m') = DATE_FORMAT(NOW() - INTERVAL 1 MONTH, '%Y-%m')
    group by
      BOS_Group_Id
    '''
    df_sa_last_month = pd.read_sql(text(sql), engine.connect())
   

     # 本月出货
    sql=f'''
    SELECT
      BOS_Group_Id,
      sum(Rebate_Sales_Amount_Cny) as this_month_sales_amount
    FROM
      ICCPP_DW_DWS.bos_sa_correct
    WHERE
      DATE_FORMAT(Stock_Out_Date, '%Y-%m') = DATE_FORMAT(NOW(), '%Y-%m')
    group by
      BOS_Group_Id
    '''
    df_sa_this_month = pd.read_sql(text(sql), engine.connect())


    # 拜访信息
    sql=f'''
    SELECT
      a.otherData,
      b.value as  visit_date
    FROM
      ICCPP_DW_DWS.crm_form a
    left join ICCPP_DW_DWS.crm_form_data b on
      a.form_id = b.form_id
      and name = 'Date of Visit'
    '''
    df_visit  = pd.read_sql(text(sql), engine.connect())

    # 将 JSON 数据展平，并拆分为多个 DataFrame
    dfs = []
    for index, row in df_visit.iterrows():
        json_data = json.loads(row['otherData'])
        temp_df = pd.json_normalize(json_data)
        temp_df['visit_date'] = row['visit_date']
        dfs.append(temp_df)
    df_visit = pd.concat(dfs, ignore_index=True)

    # 获取当前月份
    current_month = today.strftime("%Y-%m")
    # 获取上个月份
    prev_month = (today - timedelta(days=30)).strftime("%Y-%m")


    df_visit_this_month=df_visit[df_visit['visit_date'].str.contains(current_month, na=False)]
    df_visit_last_month=df_visit[df_visit['visit_date'].str.contains(prev_month, na=False)]

    df_visit_this_month_num = df_visit_this_month['customerNo'].value_counts().reset_index()
    df_visit_this_month_num.columns=['crm_code','this_month_visit_nums']

    df_visit_last_month_num = df_visit_last_month['customerNo'].value_counts().reset_index()
    df_visit_last_month_num.columns=['crm_code','last_month_visit_nums']



    df=pd.merge(df_custom,df_so_this_month,how='left',on='BOS_Group_Id')
    df=pd.merge(df,df_so_last_month,how='left',on='BOS_Group_Id')
    df=pd.merge(df,df_sa_last_month,how='left',on='BOS_Group_Id')
    df=pd.merge(df,df_sa_this_month,how='left',on='BOS_Group_Id')

    # 合并拜访信息
    df=pd.merge(df,df_visit_this_month_num,how='left',on='crm_code')
    df=pd.merge(df,df_visit_last_month_num,how='left',on='crm_code')

    
    # 缺省值
    df['Business_Region'] = df['Business_Region'].fillna('未分配区域')

    df['Sales_Representative_Name']=np.where(pd.isnull(df['Sales_Representative_Name_x']),df['Sales_Representative_Name_y'],df['Sales_Representative_Name_x'])
    
    df['last_month_order_amount'].fillna(0,inplace=True)
    df['this_month_order_amount'].fillna(0,inplace=True)
    df['last_month_sales_amount'].fillna(0,inplace=True)
    df['this_month_sales_amount'].fillna(0,inplace=True)
    df['this_month_visit_nums'].fillna(0,inplace=True)
    df['last_month_visit_nums'].fillna(0,inplace=True)

    # 指定排序顺序
    custom_order = ['美中区域', '美东区域', '美西区域', '加拿大区域', '英国区域', '法国区域', '意大利区域', '德国区域', '俄罗斯区域', '马来区域', 
    '印尼区域', '菲律宾区域', '新兴贸易区域', '澳新区域', '内贸区域','中东非区域','电商部','未分配区域']

    df_voopoo=df[df['Business_Group_Name'] == 'VOOPOO海外'].copy()

    # 按指定顺序排序
    df_voopoo['order'] = df_voopoo['Business_Region'].map({v: i for i, v in enumerate(custom_order)})
    df_voopoo = df_voopoo.sort_values(by=['order', 'Sales_Representative_Name','this_month_order_amount'],ascending=[True,True,False]).drop('order', axis=1)

   

    df_zovoo=df[df['Business_Group_Name'] == 'ZOVOO海外'].copy()
    # 按指定顺序排序
    df_zovoo['order'] = df_zovoo['Business_Region'].map({v: i for i, v in enumerate(custom_order)})
    df_zovoo = df_zovoo.sort_values(by=['order', 'Sales_Representative_Name','this_month_order_amount'],ascending=[True,True,False]).drop('order', axis=1)


    # df_odm=df[df['Business_Group_Name'] == 'ODM事业部']
    # df_odm = df_odm.sort_values(by='Business_Region')


    lo_infos['voopooRows'] = json.loads(df_voopoo.to_json(orient='records'))
    lo_infos['zovooRows'] = json.loads(df_zovoo.to_json(orient='records'))
    #lo_infos['odmRows'] = json.loads(df_odm.to_json(orient='records'))



    lo_infos['data_time'] = today.strftime(r'%Y-%m-%d %H:%M:%S')

    file_path=f'report/custom/CustomReport{yesterday_str}.xlsx'  

    lo_infos['sheet_name'] = 'sheet'   

    xlst.write2(lo_infos, file_path,'客户报告')
    print('生成',file_path)


    wb = openpyxl.load_workbook(file_path)
    sheet = wb["sheet"]

    # 先把所有内容放入一个List中
    sumList = []
    # 第几行开始
    start_row=3
    end_row=start_row+sheet.max_row
    column=3

    for i in range(start_row, end_row):
        value = sheet.cell(row=i, column=column).value
        if value:
            sumList.append(value)
        else:
            break


    # 开始合并单元格
    preRow = 0
    finRow = 0
    flag = sumList[0]
    for i in range(len(sumList)):
        if sumList[i] != flag:
            flag = sumList[i]
            finRow = i - 1
            if finRow >= preRow:
                sheet.merge_cells("C{}:C{}".format(preRow+start_row, finRow+start_row))
                sheet.merge_cells("B{}:B{}".format(preRow+start_row, finRow+start_row))
                preRow = finRow + 1
        if i == len(sumList) - 1:
            finRow = i
            sheet.merge_cells("C{}:C{}".format(preRow+start_row, finRow+start_row))
            sheet.merge_cells("B{}:B{}".format(preRow+start_row, finRow+start_row))

    
    # C列对齐
    wsArea = sheet["C{}:C{}".format(start_row, end_row)]
    for row in wsArea:
        for cell in row:    
            # 添加换行
            if cell.value:
                val='\n'+cell.value
                if cell.value=='占位符':
                  val=''
                elif cell.value=='中东非区域':
                  # 临时处理
                  val=cell.value
                cell_coordinate = cell.coordinate
                sheet[cell_coordinate]=val

            cell.alignment = Alignment(horizontal='center', vertical='top',wrapText=True)
            
    # B列对齐
    wsArea = sheet["B{}:B{}".format(start_row, end_row)]
    for row in wsArea:
        for cell in row:    
            cell.alignment = Alignment(horizontal='center', vertical='center',wrapText=True)        
    

    # 保存表
    wb.save(file_path)

   



    # # 转换成pdf
    # new_path = f'report/quality/产品品质问题数据报告{yesterday_str}.pdf'
    # password=''
    # excel=convert.excel(file_path,new_path,password)
    # excel.to_pdf()  
    # print()

    return file_path

if __name__ == '__main__':
    main()
